import uuid
import csv
import io
from datetime import datetime, timezone
from typing import List, Optional
from fastapi import APIRouter, Depends, HTTPException, status
from fastapi.responses import StreamingResponse, Response
from google.cloud import firestore
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from app.db.session import get_db
from app.core.deps import get_current_admin
from app.models.form import Form
from app.models.form_field import FormField, FieldType
from app.models.user import User
from app.models.submission import Submission, SubmissionStatus
from app.schemas.form import (
    FormCreate,
    FormUpdate,
    FormRead,
    FormDetailRead,
    FieldReorderRequest,
    FormAnalyticsResponse,
    AdminStatsResponse
)
from app.schemas.form_field import FormFieldUpdate, FormFieldRead
from app.utils.converters import dict_to_form, form_to_dict, dict_to_submission, to_datetime as _to_datetime

router = APIRouter()




@router.post("/", response_model=FormRead, status_code=status.HTTP_201_CREATED)
async def create_form(
    payload: FormCreate,
    db: firestore.AsyncClient = Depends(get_db),
    current_admin: User = Depends(get_current_admin),
) -> Form:
    """Create a new form. Slug must be unique."""
    slug_docs = await db.collection("forms").where("slug", "==", payload.slug).limit(1).get()
    if slug_docs:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Form slug is already taken.",
        )

    form = Form(
        name=payload.name,
        slug=payload.slug,
        description=payload.description,
        is_active=payload.is_active,
        unique_field_ids=payload.unique_field_ids,
    )
    await db.collection("forms").document(str(form.id)).set(form_to_dict(form))
    return form


@router.get("/", response_model=List[FormRead])
async def list_forms(
    db: firestore.AsyncClient = Depends(get_db),
    current_admin: User = Depends(get_current_admin),
) -> List[Form]:
    """List all forms, ordered by creation date descending."""
    docs = await db.collection("forms").order_by("created_at", direction=firestore.Query.DESCENDING).get()
    forms = [dict_to_form(doc.id, doc.to_dict()) for doc in docs]
    
    for form in forms:
        subs_docs = await db.collection("submissions").where("form_id", "==", str(form.id)).get()
        non_archived_subs = [d for d in subs_docs if d.to_dict().get("status") != "archived"]
        form.submission_counter = len(non_archived_subs)
        
    return forms


@router.get("/public/{slug}", response_model=FormDetailRead)
async def get_public_form(
    slug: str,
    db: firestore.AsyncClient = Depends(get_db),
) -> Form:
    """Public endpoint to fetch form details by slug. Unauthenticated."""
    slug_docs = await db.collection("forms").where("slug", "==", slug).limit(1).get()
    if not slug_docs:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Form not found.",
        )
    doc = slug_docs[0]
    form = dict_to_form(doc.id, doc.to_dict())
    if not form.is_active:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="This form is currently closed.",
        )
    return form


@router.get("/admin/stats", response_model=AdminStatsResponse)
async def get_admin_stats(
    db: firestore.AsyncClient = Depends(get_db),
    current_admin: User = Depends(get_current_admin),
) -> dict:
    """Fetch live counts for the admin dashboard (excluding archived submissions)."""
    forms_docs = await db.collection("forms").get()
    total_forms = len(forms_docs)

    subs_docs = await db.collection("submissions").get()
    non_archived_subs = [d for d in subs_docs if d.to_dict().get("status") != "archived"]
    total_submissions = len(non_archived_subs)

    return {
        "total_forms": total_forms,
        "total_submissions": total_submissions,
    }



@router.get("/{id}/analytics", response_model=FormAnalyticsResponse)
async def get_form_analytics(
    id: uuid.UUID,
    force_refresh: bool = False,
    db: firestore.AsyncClient = Depends(get_db),
    current_admin: User = Depends(get_current_admin),
):
    """Retrieve form submissions analytics. Uses pre-computed cache by default."""
    form_doc = await db.collection("forms").document(str(id)).get()
    if not form_doc.exists:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Form not found.",
        )

    cache_doc = await db.collection("analytics_cache").document(str(id)).get()

    if not cache_doc.exists or force_refresh:
        from app.services.analytics import update_analytics_cache
        from app.models.analytics_cache import AnalyticsCache
        cache = await update_analytics_cache(id, db)
    else:
        data = cache_doc.to_dict()
        from app.models.analytics_cache import AnalyticsCache
        cache = AnalyticsCache(
            id=id,
            form_id=id,
            total_submissions=data.get("total_submissions", 0),
            status_counts=data.get("status_counts", {}),
            daily_counts=data.get("daily_counts", []),
            field_stats=data.get("field_stats", {}),
            computed_at=data.get("computed_at")
        )

    today_str = datetime.now(timezone.utc).date().isoformat()
    today_submissions = 0
    for day in cache.daily_counts:
        if day.get("date") == today_str:
            today_submissions = day.get("count", 0)
            break

    return FormAnalyticsResponse(
        total_submissions=cache.total_submissions,
        today_submissions=today_submissions,
        daily_counts=cache.daily_counts,
        field_stats=cache.field_stats,
        computed_at=_to_datetime(cache.computed_at)
    )


@router.get("/{id}", response_model=FormDetailRead)
async def get_form(
    id: uuid.UUID,
    db: firestore.AsyncClient = Depends(get_db),
    current_admin: User = Depends(get_current_admin),
) -> Form:
    """Fetch form details by ID, including its fields and active submission counts."""
    doc = await db.collection("forms").document(str(id)).get()
    if not doc.exists:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Form not found.",
        )
    form = dict_to_form(doc.id, doc.to_dict())
    
    subs_docs = await db.collection("submissions").where("form_id", "==", str(form.id)).get()
    non_archived_subs = [d for d in subs_docs if d.to_dict().get("status") != "archived"]
    form.submission_counter = len(non_archived_subs)
    return form


@router.patch("/{id}", response_model=FormRead)
async def update_form(
    id: uuid.UUID,
    payload: FormUpdate,
    db: firestore.AsyncClient = Depends(get_db),
    current_admin: User = Depends(get_current_admin),
) -> Form:
    """Update form metadata/settings."""
    doc_ref = db.collection("forms").document(str(id))
    doc = await doc_ref.get()
    if not doc.exists:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Form not found.",
        )
    form = dict_to_form(doc.id, doc.to_dict())

    if payload.slug and payload.slug != form.slug:
        slug_docs = await db.collection("forms").where("slug", "==", payload.slug).limit(1).get()
        if slug_docs:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Form slug is already taken.",
            )

    update_data = payload.model_dump(exclude_unset=True)
    for key, val in update_data.items():
        setattr(form, key, val)

    form.updated_at = datetime.now(timezone.utc)
    await doc_ref.set(form_to_dict(form))
    return form


@router.delete("/{id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_form(
    id: uuid.UUID,
    db: firestore.AsyncClient = Depends(get_db),
    current_admin: User = Depends(get_current_admin),
) -> None:
    """Delete form. Automatically cascades to fields and submissions."""
    doc_ref = db.collection("forms").document(str(id))
    doc = await doc_ref.get()
    if not doc.exists:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Form not found.",
        )
    
    # Use a write batch to group deletions into a single network request
    batch = db.batch()
    batch.delete(doc_ref)

    subs_docs = await db.collection("submissions").where("form_id", "==", str(id)).get()
    for sub in subs_docs:
        batch.delete(db.collection("submissions").document(sub.id))

    batch.delete(db.collection("analytics_cache").document(str(id)))
    await batch.commit()



@router.post("/{id}/fields", response_model=List[FormFieldRead])
async def sync_form_fields(
    id: uuid.UUID,
    payload: List[FormFieldUpdate],
    db: firestore.AsyncClient = Depends(get_db),
    current_admin: User = Depends(get_current_admin),
) -> List[FormField]:
    """Sync fields for a form. 
    Adds new fields, updates existing, and deletes omitted fields.
    """
    doc_ref = db.collection("forms").document(str(id))
    doc = await doc_ref.get()
    if not doc.exists:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Form not found.",
        )
    form = dict_to_form(doc.id, doc.to_dict())

    updated_fields = []
    for item in payload:
        f_id = item.id or uuid.uuid4()
        existing_field = next((f for f in form.fields if f.id == f_id), None)
        if existing_field:
            existing_field.field_type = FieldType(item.field_type)
            existing_field.label = item.label
            existing_field.placeholder = item.placeholder
            existing_field.description = item.description
            existing_field.default_value = item.default_value
            existing_field.is_required = item.is_required
            existing_field.order = item.order
            existing_field.options = item.options
            existing_field.conditions = item.conditions
            existing_field.file_accepted_types = item.file_accepted_types
            existing_field.file_max_size_mb = item.file_max_size_mb
            existing_field.file_max_count = item.file_max_count
            existing_field.updated_at = datetime.now(timezone.utc)
            updated_fields.append(existing_field)
        else:
            new_field = FormField(
                id=f_id,
                form_id=id,
                field_type=FieldType(item.field_type),
                label=item.label,
                placeholder=item.placeholder,
                description=item.description,
                default_value=item.default_value,
                is_required=item.is_required,
                order=item.order,
                options=item.options,
                conditions=item.conditions,
                file_accepted_types=item.file_accepted_types,
                file_max_size_mb=item.file_max_size_mb,
                file_max_count=item.file_max_count,
            )
            updated_fields.append(new_field)

    updated_fields.sort(key=lambda x: x.order)
    form.fields = updated_fields
    form.updated_at = datetime.now(timezone.utc)

    await doc_ref.set(form_to_dict(form))
    return form.fields


@router.put("/{id}/fields/reorder", status_code=status.HTTP_204_NO_CONTENT)
async def reorder_form_fields(
    id: uuid.UUID,
    payload: FieldReorderRequest,
    db: firestore.AsyncClient = Depends(get_db),
    current_admin: User = Depends(get_current_admin),
) -> None:
    """Update field orders in bulk based on a list of field IDs."""
    doc_ref = db.collection("forms").document(str(id))
    doc = await doc_ref.get()
    if not doc.exists:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Form not found.",
        )
    form = dict_to_form(doc.id, doc.to_dict())

    fields_lookup = {f.id: f for f in form.fields}
    for idx, field_id in enumerate(payload.field_ids):
        if field_id in fields_lookup:
            fields_lookup[field_id].order = idx

    form.fields.sort(key=lambda x: x.order)
    form.updated_at = datetime.now(timezone.utc)
    await doc_ref.set(form_to_dict(form))


async def fetch_submissions_for_export(form_id: uuid.UUID, db: firestore.AsyncClient):
    form_doc = await db.collection("forms").document(str(form_id)).get()
    if not form_doc.exists:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Form not found."
        )
    form = dict_to_form(form_doc.id, form_doc.to_dict())
    fields = form.fields

    subs_docs = await db.collection("submissions").where("form_id", "==", str(form_id)).get()
    submissions = [dict_to_submission(doc.id, doc.to_dict()) for doc in subs_docs]
    submissions.sort(key=lambda x: x.submitted_at, reverse=True)
    return fields, submissions


@router.get("/{id}/export/csv")
async def export_submissions_csv(
    id: uuid.UUID,
    db: firestore.AsyncClient = Depends(get_db),
    current_admin: User = Depends(get_current_admin),
):
    """Export all submissions for a form in CSV format."""
    form_res = await db.collection("forms").document(str(id)).get()
    if not form_res.exists:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Form not found.",
        )
    form = dict_to_form(form_res.id, form_res.to_dict())

    fields, submissions = await fetch_submissions_for_export(id, db)

    output = io.StringIO()
    writer = csv.writer(output)

    header = ["Submission ID", "Status", "Submitted At"]
    for field in fields:
        header.append(field.label)
    writer.writerow(header)

    for sub in submissions:
        row = [
            sub.submission_id,
            sub.status.value if hasattr(sub.status, "value") else str(sub.status),
            sub.submitted_at.strftime("%Y-%m-%d %H:%M:%S")
        ]
        for field in fields:
            if field.field_type == FieldType.FILE:
                files = [f for f in sub.file_uploads if f.field_id == field.id]
                row.append(", ".join(f.cloudinary_secure_url for f in files))
            else:
                sub_val = next((v for v in sub.values if v.field_id == field.id), None)
                if sub_val:
                    if sub_val.value_json is not None:
                        if isinstance(sub_val.value_json, list):
                            row.append(", ".join(str(x) for x in sub_val.value_json))
                        else:
                            row.append(str(sub_val.value_json))
                    elif sub_val.value_text is not None:
                        row.append(sub_val.value_text)
                    else:
                        row.append("")
                else:
                    row.append("")
        writer.writerow(row)

    output.seek(0)
    filename = f"export_{form.slug}_{datetime.now().strftime('%Y%m%d%H%M%S')}.csv"
    return StreamingResponse(
        io.StringIO(output.getvalue()),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/{id}/export/xlsx")
async def export_submissions_xlsx(
    id: uuid.UUID,
    db: firestore.AsyncClient = Depends(get_db),
    current_admin: User = Depends(get_current_admin),
):
    """Export all submissions for a form in Excel (XLSX) format."""
    form_res = await db.collection("forms").document(str(id)).get()
    if not form_res.exists:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Form not found.",
        )
    form = dict_to_form(form_res.id, form_res.to_dict())

    fields, submissions = await fetch_submissions_for_export(id, db)

    wb = Workbook()
    ws = wb.active
    ws.title = "Submissions"

    header = ["Submission ID", "Status", "Submitted At"]
    for field in fields:
        header.append(field.label)
    ws.append(header)

    for col_idx in range(1, len(header) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = Font(bold=True)

    for sub in submissions:
        row = [
            sub.submission_id,
            sub.status.value if hasattr(sub.status, "value") else str(sub.status),
            sub.submitted_at.strftime("%Y-%m-%d %H:%M:%S")
        ]
        for field in fields:
            if field.field_type == FieldType.FILE:
                files = [f for f in sub.file_uploads if f.field_id == field.id]
                row.append(", ".join(f.cloudinary_secure_url for f in files))
            else:
                sub_val = next((v for v in sub.values if v.field_id == field.id), None)
                if sub_val:
                    if sub_val.value_json is not None:
                        if isinstance(sub_val.value_json, list):
                            row.append(", ".join(str(x) for x in sub_val.value_json))
                        else:
                            row.append(str(sub_val.value_json))
                    elif sub_val.value_text is not None:
                        row.append(sub_val.value_text)
                    else:
                        row.append("")
                else:
                    row.append("")
        ws.append(row)

    for col in ws.columns:
        max_len = 0
        for cell in col:
            val_str = str(cell.value or "")
            if len(val_str) > max_len:
                max_len = len(val_str)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 10)

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)

    filename = f"export_{form.slug}_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
    return Response(
        content=out.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
